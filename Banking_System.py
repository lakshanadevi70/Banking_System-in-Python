
from datetime import datetime
from getpass import getpass
from openpyxl import load_workbook

EXCEL_FILE = "Book.xlsx"
SHEET_USERS = "Sheet1"   # Credentials
SHEET_TXNS = "Sheet2"    # Transactions log

def load_users():
    """Return a dict {user_id: password} from Sheet1."""
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_USERS]
    users = {}
    for uid, pwd in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if uid and pwd:
            users[str(uid).strip()] = str(pwd).strip()
    return users

def log_transaction(user, action, deposit="", withdraw=""):
    """Append a row to Sheet2 with timestamp."""
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_TXNS]
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([user, action, deposit, withdraw, ts])
    wb.save(EXCEL_FILE)

def main():
    users = load_users()

    print("=== Welcome to Secure Banking System ===")
    user_id = input("Enter User ID: ").strip()
    password = getpass("Enter Password: ").strip()

    if not (user_id in users and users[user_id] == password):
        print("❌ Invalid User ID or Password! Access Denied.")
        return

    print("\nLogin Successful! ✅")
    balance = 1000  # session balance only

    while True:
        print("\n🏦 Banking System")
        print("1. Check Balance")
        print("2. Deposit")
        print("3. Withdraw")
        print("4. Exit")

        try:
            choice = int(input("Enter your choice: ")) 
        except ValueError:
            print("Please enter a valid number (1-4).")
            continue

        if choice == 1:
            print("Current Balance:", balance)
            log_transaction(user_id, "Check Balance")

        elif choice == 2:
            try:
                amount = int(input("Enter deposit amount: "))
                if amount > 0:
                    balance += amount
                    print("Amount Deposited. New Balance:", balance)
                    log_transaction(user_id, "Deposit", deposit=amount)
                else:
                    print("Enter a positive amount.")
            except ValueError:
                print("Please enter a valid number.")

        elif choice == 3:
            try:
                amount = int(input("Enter withdrawal amount: "))
                if amount <= 0:
                    print("Enter a positive amount.")
                elif amount <= balance:
                    balance -= amount
                    print("Amount Withdrawn. New Balance:", balance)
                    log_transaction(user_id, "Withdraw", withdraw=amount)
                else:
                    print("Insufficient balance!")
            except ValueError:
                print("Please enter a valid number.")

        elif choice == 4:
            print("Exiting... Thank you for banking with us! 🙏")
            log_transaction(user_id, "Exit")
            break

        else:
            print("Invalid choice! Please try again.")

if __name__ == "__main__":
    main()
